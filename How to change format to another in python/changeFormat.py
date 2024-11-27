import openpyxl as op
# retrieves the data from the excel format
data = op.load_workbook(r"C:\Users\SondreHaugom\OneDrive - Telemark fylkeskommune\Sondre\Prakis oppgaver høst 2024\endreFromat\innbyggerliste-skien-test-20.11.2024.xlsx")
sheets =  data.sheetnames
sheet = data.active

# creates a new Excel file for export
ny_fil = op.Workbook()
ny_sheet = ny_fil.active
ny_sheet.tittle = "New format with more tabels"

# fetches columns
ny_sheet.append(["Fornavn","Etternavn","Adresse","Post nummer","Post sted","Identifikasjonsnummer_foedselsEllerDNummer"])

# finds max number of rows
rows = sheet.max_row

# retrieves the information from the first excel sheet
for i in range(2, rows + 1):
    fornavn = sheet[f"D{i}"].value
    etternavn = sheet[f"E{i}"].value
    adresse = sheet[f"F{i}"].value
    post_nummer = sheet[f"G{i}"].value
    post_sted = sheet[f"H{i}"].value
    Identifikasjonsnummer_foedselsEllerDNummer = sheet[f"I{i}"].value


    if fornavn and etternavn and adresse and post_nummer and post_sted and Identifikasjonsnummer_foedselsEllerDNummer:

        ny_sheet.append([fornavn,etternavn,adresse,post_nummer,post_sted,Identifikasjonsnummer_foedselsEllerDNummer])

# create the new excel file
output_path = r"C:\Users\SondreHaugom\OneDrive - Telemark fylkeskommune\Sondre\Prakis oppgaver høst 2024\endreFromat\New format with more tabels.xlsx"
ny_fil.save(output_path)

print(f"Data ekspoterer til {output_path}")